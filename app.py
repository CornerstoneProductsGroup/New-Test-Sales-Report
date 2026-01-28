import os

def _fmt_units(x):
    try:
        return f"{float(x):,.0f}"
    except Exception:
        return ""

def _fmt_money(x):
    try:
        return f"${float(x):,.2f}"
    except Exception:
        return ""

# Persistent storage file (NOTE: Streamlit Cloud may wipe local files on cold restart; use Backup/Restore tab).
DB_FILE = os.path.join(os.path.dirname(__file__), 'app.db')


def _file_md5(path: str) -> str:
    h = hashlib.md5()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


BIGGER_TABLES_CSS = """
<style>
div[data-testid="stDataFrame"] table td, div[data-testid="stDataFrame"] table th,
div[data-testid="stDataEditor"] table td, div[data-testid="stDataEditor"] table th {
  font-size: 11px !important;
  padding: 0.12rem 0.22rem !important;
  white-space: nowrap;
}
</style>
"""

# ---------- AUTOLOAD_VENDOR_MAP ----------
# Load Vendor-SKU pricing map automatically from local file on startup
def autoload_vendor_map(conn):
    import pandas as pd, os
    map_path = os.path.join(os.path.dirname(__file__), "Vendor-SKU Map.xlsx")
    if not os.path.exists(map_path):
        return
    df = pd.read_excel(map_path)
    df.columns = [c.strip().lower() for c in df.columns]
    df = df.rename(columns={
        "retailer": "retailer",
        "vendor": "vendor",
        "sku": "sku",
        "price": "unit_price",
        "unit price": "unit_price",
    })
    df = df[["retailer","vendor","sku","unit_price"]].dropna()
    df["retailer"] = df["retailer"].astype(str).str.strip()
    df["vendor"] = df["vendor"].astype(str).str.strip()
    df["sku"] = df["sku"].astype(str).str.strip()
    df["unit_price"] = pd.to_numeric(df["unit_price"], errors="coerce").fillna(0)

    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS sku_mapping (
            retailer TEXT,
            vendor TEXT,
            sku TEXT,
            unit_price REAL,
            UNIQUE(retailer, vendor, sku)
        )
    """)
    for _,r in df.iterrows():
        cur.execute(
            "INSERT OR REPLACE INTO sku_mapping (retailer,vendor,sku,unit_price) VALUES (?,?,?,?)",
            (r["retailer"], r["vendor"], r["sku"], float(r["unit_price"]))
        )
    conn.commit()
# ---------- END AUTOLOAD_VENDOR_MAP ----------

# ---------- PRICE + VENDOR MAP HELPERS (added) ----------
def try_load_prices(conn):
    """Return a dataframe with columns: retailer, sku, price (unit_price)."""
    try:
        df = pd.read_sql_query("SELECT retailer, sku, unit_price FROM sku_mapping", conn)
    except Exception:
        return pd.DataFrame(columns=["retailer","sku","price"])
    df = df.copy()
    df["retailer"] = df["retailer"].astype(str).str.strip()
    df["sku"] = df["sku"].astype(str).str.strip()
    df["price"] = pd.to_numeric(df["unit_price"], errors="coerce")
    return df[["retailer","sku","price"]]

def load_vendor_map():
    """Return a dataframe with columns: sku, vendor using sku_mapping table."""
    try:
        df = pd.read_sql_query("SELECT sku, vendor FROM sku_mapping", conn)
    except Exception:
        # fallback to file if table missing
        try:
            import pandas as pd, os
            map_path = os.path.join(os.path.dirname(__file__), "Vendor-SKU Map.xlsx")
            if not os.path.exists(map_path):
                return pd.DataFrame(columns=["sku","vendor"])
            d = pd.read_excel(map_path)
            d.columns = [c.strip().lower() for c in d.columns]
            if "sku" not in d.columns or "vendor" not in d.columns:
                return pd.DataFrame(columns=["sku","vendor"])
            d = d[["sku","vendor"]].dropna()
            d["sku"] = d["sku"].astype(str).str.strip()
            d["vendor"] = d["vendor"].astype(str).str.strip()
            return d
        except Exception:
            return pd.DataFrame(columns=["sku","vendor"])
    df = df.dropna().copy()
    df["sku"] = df["sku"].astype(str).str.strip()
    df["vendor"] = df["vendor"].astype(str).str.strip()
    # If a SKU appears for multiple retailers, keep the most common vendor label
    df = df.groupby("sku", as_index=False)["vendor"].agg(lambda s: s.value_counts().index[0] if len(s.dropna()) else "Unknown")
    return df
# ---------- END HELPERS ----------



COMPACT_TABLE_CSS = """
<style>
/* Reduce padding inside Streamlit dataframes */
div[data-testid="stDataFrame"] table td,
div[data-testid="stDataFrame"] table th {
  padding: 0.25rem 0.4rem !important;
  white-space: nowrap;
}

/* Prevent tables from stretching wide */
div[data-testid="stDataFrame"] {
  max-width: 600px;
}
</style>
"""

SPLIT_TABLE_CSS = """
<style>
/* Slightly reduce spacing between Streamlit columns */
div[data-testid="stHorizontalBlock"] { gap: 0.5rem !important; }
</style>
"""
def init_meta(conn):
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS app_meta (
            k TEXT PRIMARY KEY,
            v TEXT
        )
        """
    )
    conn.commit()

def get_meta(conn, k: str):
    init_meta(conn)
    row = conn.execute("SELECT v FROM app_meta WHERE k = ?", (k,)).fetchone()
    return row[0] if row else None

def set_meta(conn, k: str, v: str):
    init_meta(conn)
    conn.execute("INSERT INTO app_meta(k, v) VALUES(?, ?) ON CONFLICT(k) DO UPDATE SET v=excluded.v", (k, v))
    conn.commit()

def ensure_mapping_loaded(conn, mapping_path: str):
    """
    Ensures sku_mapping is populated from the bundled vendor map Excel.
    Reloads automatically if the file changes.
    """
    init_meta(conn)
    file_hash = _file_md5(mapping_path) if os.path.exists(mapping_path) else None
    prev_hash = get_meta(conn, "mapping_hash")

    # If table empty OR file changed, import
    try:
        cnt = conn.execute("SELECT COUNT(*) FROM sku_mapping WHERE active = 1").fetchone()[0]
    except Exception:
        cnt = 0

    if (cnt == 0) or (file_hash and prev_hash != file_hash):
        # Re-import mapping (uses existing import function if present)
        df_map = pd.read_excel(mapping_path)
        # Normalize expected columns
        df_map.columns = [str(c).strip() for c in df_map.columns]
        # Try common names
        col_retailer = next((c for c in df_map.columns if c.lower() in ["retailer", "store"]), None)
        col_vendor = next((c for c in df_map.columns if c.lower() in ["vendor", "manufacturer"]), None)
        col_sku = next((c for c in df_map.columns if c.lower() in ["sku", "vendor sku", "retailer sku"]), None)
        col_price = next((c for c in df_map.columns if "price" in c.lower()), None)

        if not (col_retailer and col_vendor and col_sku):
            raise ValueError("Vendor map is missing required columns (Retailer, Vendor, SKU).")

        df_norm = pd.DataFrame({
            "retailer": df_map[col_retailer].astype(str).str.strip(),
            "vendor": df_map[col_vendor].astype(str).str.strip(),
            "sku": df_map[col_sku].astype(str).str.strip(),
        })
        if col_price:
            df_norm["unit_price"] = pd.to_numeric(df_map[col_price], errors="coerce")
        else:
            df_norm["unit_price"] = pd.NA

        df_norm = df_norm[df_norm["sku"].ne("") & df_norm["retailer"].ne("")].copy()
        df_norm["active"] = 1
        df_norm["sort_order"] = pd.NA

        # De-duplicate to avoid UNIQUE/PK conflicts (keep first)
        df_norm = df_norm.drop_duplicates(subset=["retailer", "sku"], keep="first")

        # Replace existing mapping table atomically
        df_norm.to_sql("sku_mapping", conn, if_exists="replace", index=False)
        conn.commit()
        conn.execute("CREATE INDEX IF NOT EXISTS idx_sku_mapping_retailer_sku ON sku_mapping(retailer, sku)")
        conn.commit()
        if file_hash:
            set_meta(conn, "mapping_hash", file_hash)
import streamlit as st
import pandas as pd
import hashlib
import sqlite3
import json
from datetime import date, datetime, timedelta
from openpyxl import load_workbook
import re


def render_backup_restore_tab(db_file: str, map_filename: str = "Vendor-SKU Map.xlsx") -> None:
    """Render the Backup / Restore tab.

    - Download current sqlite DB (db_file) and optional vendor map.
    - Restore from uploaded .db or .zip (containing a .db and optional vendor map).

    Designed to be safe:
      - creates a timestamped .bak copy before overwriting the DB
      - validates restored DB is readable by sqlite
    """
    import io
    import zipfile
    import shutil
    import tempfile
    from datetime import datetime
    from pathlib import Path

    st.header("Backup / Restore")
    st.caption("Download your current database (and optionally the Vendor-SKU map) so you can restore it if the app cold-restarts or you move environments.")

    app_dir = Path(__file__).resolve().parent
    map_path = app_dir / map_filename

    # ---------- Current file status ----------
    st.subheader("Current local files")
    c1, c2, c3 = st.columns(3)

    with c1:
        st.write("Database file")
        if os.path.exists(db_file):
            st.success("Found")
            st.code(db_file)
        else:
            st.warning("Not found yet")

    with c2:
        st.write("Vendor-SKU Map")
        if map_path.exists():
            st.success("Found")
            st.code(str(map_path))
        else:
            st.warning("Not found")

    with c3:
        st.write("Notes")
        st.write("Restoring overwrites local files in this app instance.")

    # ---------- Download backup ----------
    st.subheader("Create backup (download)")
    include_map = st.checkbox("Include Vendor-SKU Map.xlsx in backup zip", value=True)

    if os.path.exists(db_file):
        with open(db_file, "rb") as f:
            db_bytes = f.read()

        st.download_button(
            "Download database only (.db)",
            data=db_bytes,
            file_name=f"{Path(db_file).stem}_{datetime.now().strftime('%Y-%m-%d')}.db",
            mime="application/octet-stream",
            use_container_width=True,
        )

        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as z:
            z.writestr(Path(db_file).name, db_bytes)
            if include_map and map_path.exists():
                z.write(str(map_path), arcname=map_filename)

        st.download_button(
            "Download backup zip (.zip)",
            data=buf.getvalue(),
            file_name=f"sales_app_backup_{datetime.now().strftime('%Y-%m-%d_%H%M')}.zip",
            mime="application/zip",
            use_container_width=True,
        )

        try:
            st.caption(f"DB size: {len(db_bytes):,} bytes • MD5: {_file_md5(db_file)}")
        except Exception:
            pass
    else:
        st.info("No database found yet. Upload a week workbook first (or restore a backup) to create the local database.")

    st.divider()

    # ---------- Restore ----------
    st.subheader("Restore from backup (upload)")
    st.warning("Restoring overwrites the current local database (and map file if included). Download a backup first if you need the current state.")

    upload = st.file_uploader(
        "Upload a backup file",
        type=["zip", "db"],
        help="Upload either a .db file or a .zip that contains a .db (and optionally Vendor-SKU Map.xlsx).",
    )

    if upload is None:
        return

    filename = (upload.name or "").lower()
    data = upload.getvalue()

    if st.button("Restore now", type="primary", use_container_width=True):
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")

        # Safety copy of existing DB
        if os.path.exists(db_file):
            try:
                shutil.copy2(db_file, db_file + f".bak_{ts}")
            except Exception:
                pass

        try:
            if filename.endswith(".db"):
                with open(db_file, "wb") as f:
                    f.write(data)

            elif filename.endswith(".zip"):
                tmp = tempfile.mkdtemp(prefix="restore_")
                try:
                    with zipfile.ZipFile(io.BytesIO(data), "r") as z:
                        names = z.namelist()

                        # Pick the first .db we find
                        db_candidates = [n for n in names if n.lower().endswith(".db")]
                        if not db_candidates:
                            raise ValueError("No .db file found inside the zip.")
                        db_name = db_candidates[0]

                        z.extract(db_name, tmp)
                        extracted_db = os.path.join(tmp, db_name)
                        os.makedirs(os.path.dirname(db_file) or ".", exist_ok=True)
                        shutil.copy2(extracted_db, db_file)

                        # Optional map restore
                        map_candidates = [n for n in names if n.lower().endswith(map_filename.lower())]
                        if map_candidates:
                            map_name = map_candidates[0]
                            z.extract(map_name, tmp)
                            extracted_map = os.path.join(tmp, map_name)
                            shutil.copy2(extracted_map, str(map_path))
                finally:
                    shutil.rmtree(tmp, ignore_errors=True)

            else:
                raise ValueError("Unsupported file type. Please upload a .db or .zip.")

            # Sanity check that sqlite file is readable
            conn_test = sqlite3.connect(db_file)
            conn_test.execute("SELECT name FROM sqlite_master LIMIT 1;")
            conn_test.close()

            st.success("Restore complete. Reloading the app now…")
            st.rerun()

        except Exception as e:
            st.error(f"Restore failed: {e}")

def parse_week_label_from_filename(filename: str):
    """
    Accepts filenames like:
      'APP 1-1 thru 1-2.xlsx'
      'APP 1-5 thru 1-9.xlsx'
    Returns (label, start_date, end_date) for 2026, or (None,None,None) if not matched.
    """
    base = os.path.basename(filename)
    m = re.search(r'APP\s+(\d{1,2})-(\d{1,2})\s+thru\s+(\d{1,2})-(\d{1,2})', base, re.IGNORECASE)
    if not m:
        return None, None, None
    m1, d1, m2, d2 = map(int, m.groups())
    try:
        start = date(2026, m1, d1)
        end = date(2026, m2, d2)
    except Exception:
        return None, None, None
    label = f"{m1}-{d1} / {m2}-{d2}"
    return label, start, end

def norm_name(s: str):
    if s is None:
        return ""
    return re.sub(r"[^a-z0-9]+", "", str(s).lower())

def normalize_retailer_sheet_name(sheet_name: str):
    """
    Maps uploaded sheet names to your retailer names used in the mapping.
    Adjust here if your mapping uses different naming.
    """
    n = norm_name(sheet_name)
    if n in ("depot", "homedepot", "thehomedepot"):
        return "Depot"
    if n in ("lowes", "lowesinc"):
        return "Lowe's"
    if n == "amazon":
        return "Amazon"
    if n in ("tractorsupply", "tsc", "tractorsupplyco"):
        return "Tractor Supply"
    if n in ("depotso", "depotspecialorders", "specialorders"):
        return "Depot SO"
    return sheet_name  # fallback

def parse_app_aggregated_sheet(ws):
    """
    APP workbook format: SKU in col A, Units in col B (already aggregated).
    Skips title rows like 'Depot '.
    """
    out = {}
    for r in range(1, ws.max_row + 1):
        sku = ws.cell(r, 1).value
        qty = ws.cell(r, 2).value
        if sku is None:
            continue
        sku_str = str(sku).strip()
        if sku_str == "" or sku_str.lower() in ("sku", "vendor style"):
            continue
        # skip title row like "Depot "
        if qty is None and len(sku_str) <= 20 and sku_str.lower().strip() in ("depot", "depot ", "lowe's", "amazon", "tractor supply", "depot so"):
            continue
        try:
            q = float(qty)
        except Exception:
            continue
        # keep zeros too? We'll ignore zeros to avoid clutter
        if q == 0:
            continue
        out[sku_str] = out.get(sku_str, 0.0) + q
    return out
from pathlib import Path

APP_TITLE = "Weekly Retailer Report (Multi-week View)"
# DB_FILE is defined at top
APP_DIR = Path(__file__).resolve().parent

# -----------------------------
# Week selector (2026 only for now)
# -----------------------------
def weeks_2026():
    rows = []
    # Special partial week
    rows.append((date(2026, 1, 1), date(2026, 1, 2), "1-1 / 1-2"))
    monday = date(2026, 1, 5)
    for i in range(0, 60):
        start = monday + timedelta(weeks=i)
        end = start + timedelta(days=4)
        if start.year != 2026:
            break
        if end.year != 2026:
            end = date(2026, 12, 31)
        rows.append((start, end, f"{start.month}-{start.day} / {end.month}-{end.day}"))
        if end == date(2026, 12, 31):
            break
    return rows


MONTH_NAMES = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]

def weeks_for_months(week_meta, months_selected):
    if not months_selected:
        return []
    labels = []
    for start, _, lbl in week_meta:
        if start.month in months_selected:
            labels.append(lbl)
    return labels


def fmt_currency_str(x):
    """Format a number like $3,234.00 and negatives like ($1,174.95)."""
    try:
        if x is None or (isinstance(x, float) and pd.isna(x)):
            return ""
        v = float(x)
    except Exception:
        return ""
    s = f"${abs(v):,.2f}"
    return f"({s})" if v < 0 else s


def filter_df_year(df: pd.DataFrame, year_filter: str, col: str = "week_start") -> pd.DataFrame:
    """Filter a dataframe to a given year based on an ISO date column (YYYY-...)."""
    if df is None or df.empty or not year_filter or year_filter == "All" or col not in df.columns:
        return df
    s = df[col].astype(str)
    return df[s.str.startswith(str(year_filter))].copy()

# -----------------------------
# DB helpers
# -----------------------------
def get_conn():
    conn = sqlite3.connect(DB_FILE, check_same_thread=False)
    conn.execute("PRAGMA foreign_keys = ON;")
    return conn


def table_exists(conn, table_name: str) -> bool:
    try:
        cur = conn.cursor()
        cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table_name,))
        return cur.fetchone() is not None
    except Exception:
        return False

def init_db(conn: sqlite3.Connection):
    conn.executescript("""
    CREATE TABLE IF NOT EXISTS sku_mapping (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        retailer TEXT NOT NULL,
        vendor TEXT NOT NULL,
        sku TEXT NOT NULL,
        unit_price REAL,
        active INTEGER NOT NULL DEFAULT 1,
        sort_order INTEGER,
        UNIQUE(retailer, sku)
    );

    CREATE TABLE IF NOT EXISTS weekly_results (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        week_start TEXT NOT NULL,
        week_end TEXT NOT NULL,
        retailer TEXT NOT NULL,
        sku TEXT NOT NULL,
        units_auto REAL,
        units_override REAL,
        sales_manual REAL,
        notes TEXT,
        updated_at TEXT NOT NULL,
        UNIQUE(week_start, retailer, sku)
    );

    CREATE INDEX IF NOT EXISTS idx_weekly_results_week_retailer
    ON weekly_results(week_start, retailer);

    CREATE TABLE IF NOT EXISTS ui_state (
        key TEXT PRIMARY KEY,
        value TEXT
    );
    """)
    # In case an older DB exists, try to add unit_price (no-op if already present)
    try:
        conn.execute("ALTER TABLE sku_mapping ADD COLUMN unit_price REAL;")
    except Exception:
        pass
    conn.commit()

def mapping_count(conn):
    df = pd.read_sql_query("SELECT COUNT(*) AS n FROM sku_mapping", conn)
    return int(df.loc[0, "n"]) if not df.empty else 0

def get_ui_state(conn, key: str, default=None):
    try:
        df = pd.read_sql_query("SELECT value FROM ui_state WHERE key = ?", conn, params=(key,))
        if df.empty:
            return default
        return json.loads(df.loc[0, "value"])
    except Exception:
        return default

def set_ui_state(conn, key: str, value):
    try:
        conn.execute(
            "INSERT INTO ui_state(key, value) VALUES(?, ?) ON CONFLICT(key) DO UPDATE SET value=excluded.value",
            (key, json.dumps(value))
        )
        conn.commit()
    except Exception:
        pass

def mapping_has_any_price(conn) -> bool:
    try:
        df = pd.read_sql_query("SELECT COUNT(*) AS n FROM sku_mapping WHERE unit_price IS NOT NULL", conn)
        return int(df.loc[0, "n"]) > 0
    except Exception:
        return False

def refresh_mapping_from_bundled_if_needed(conn):
    """
    If mapping exists but has no prices populated, reload from bundled Vendor-SKU Map.xlsx.
    This fixes the common case where the DB was bootstrapped from an older map without price.
    """
    if mapping_count(conn) == 0:
        return False
    if mapping_has_any_price(conn):
        return False
    # Try reading bundled map relative to app.py
    candidates = [
        APP_DIR / "Vendor-SKU Map.xlsx",
        Path("Vendor-SKU Map.xlsx"),
    ]
    for p in candidates:
        try:
            if p.exists():
                df_map = pd.read_excel(p, sheet_name=0)
                # only reload if the file actually contains a price column
                price_col = next((c for c in df_map.columns if "price" in str(c).lower()), None)
                if price_col:
                    upsert_mapping(conn, df_map)
                    return True
        except Exception:
            continue
    return False

def upsert_mapping(conn, df: pd.DataFrame):
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]

    required = {"Retailer", "SKU", "Vendor"}
    if not required.issubset(set(df.columns)):
        raise ValueError(f"Mapping must contain columns: {sorted(required)}. Found: {list(df.columns)}")

    price_col = next((c for c in df.columns if "price" in str(c).lower()), None)

    df = df[list(required) + ([price_col] if price_col else [])].dropna(subset=["Retailer", "SKU", "Vendor"])
    df["Retailer"] = df["Retailer"].astype(str).str.strip()
    df["SKU"] = df["SKU"].astype(str).str.strip()
    df["Vendor"] = df["Vendor"].astype(str).str.strip()

    if price_col:
        # coerce to numeric
        df[price_col] = pd.to_numeric(df[price_col], errors="coerce")

    retailers = sorted(df["Retailer"].unique().tolist())
    cur = conn.cursor()
    cur.executemany("DELETE FROM sku_mapping WHERE retailer = ?", [(r,) for r in retailers])

    rows = []
    for r in retailers:
        sub = df[df["Retailer"] == r].reset_index(drop=True)
        for i, row in sub.iterrows():
            price = float(row[price_col]) if price_col and pd.notna(row[price_col]) else None
            rows.append((row["Retailer"], row["Vendor"], row["SKU"], price, 1, i + 1))

    cur.executemany("""
        INSERT INTO sku_mapping(retailer, vendor, sku, unit_price, active, sort_order)
        VALUES(?,?,?,?,?,?)
    """, rows)
    conn.commit()

def bootstrap_default_mapping(conn):
    """
    Load a bundled mapping file shipped with the app on first run (or when DB is empty).
    Works on Streamlit Cloud where the working directory can vary.
    """
    if mapping_count(conn) > 0:
        return False

    candidates = [
        APP_DIR / "Vendor-SKU Map.xlsx",
        APP_DIR / "Vendor-SKU Map - example.xlsx",
        Path("Vendor-SKU Map.xlsx"),
        Path("Vendor-SKU Map - example.xlsx"),
    ]
    for p in candidates:
        try:
            if p.exists():
                df_map = pd.read_excel(p, sheet_name=0)
                upsert_mapping(conn, df_map)
                return True
        except Exception:
            continue
    return False

    candidates = [
        APP_DIR / "Vendor-SKU Map.xlsx",
        APP_DIR / "Vendor-SKU Map - example.xlsx",
        Path("Vendor-SKU Map.xlsx"),
        Path("Vendor-SKU Map - example.xlsx"),
    ]
    for p in candidates:
        try:
            if p.exists():
                df_map = pd.read_excel(p, sheet_name=0)
                upsert_mapping(conn, df_map)
                return True
        except Exception:
            continue
    return False
    for fn in ["Vendor-SKU Map.xlsx", "Vendor-SKU Map - example.xlsx"]:
        try:
            df_map = pd.read_excel(fn, sheet_name=0)
            upsert_mapping(conn, df_map)
            return True
        except Exception:
            continue
    return False


def get_available_years(conn):
    """Return sorted list of years present in weekly_results.week_start (YYYY)."""
    try:
        df = pd.read_sql_query(
            "SELECT DISTINCT substr(week_start,1,4) AS y FROM weekly_results WHERE week_start IS NOT NULL ORDER BY y",
            conn
        )
        years = [str(y) for y in df["y"].dropna().tolist() if str(y).isdigit()]
        return years
    except Exception:
        return []

def get_retailers(conn):
    df = pd.read_sql_query("""
        SELECT DISTINCT retailer FROM sku_mapping
        WHERE active = 1
        ORDER BY retailer
    """, conn)
    return df["retailer"].tolist()

def get_mapping_for_retailer(conn, retailer: str):
    return pd.read_sql_query("""
        SELECT vendor, sku, unit_price, sort_order
        FROM sku_mapping
        WHERE active = 1 AND retailer = ?
        ORDER BY COALESCE(sort_order, 999999), vendor, sku
    """, conn, params=(retailer,))

def get_week_records(conn, retailer: str, week_starts: list[str]):
    if not week_starts:
        return pd.DataFrame(columns=["week_start","sku","units_auto","units_override","sales_manual","notes"])
    placeholders = ",".join(["?"] * len(week_starts))
    q = f"""
        SELECT week_start, sku, units_auto, units_override, sales_manual, notes
        FROM weekly_results
        WHERE retailer = ? AND week_start IN ({placeholders})
    """
    return pd.read_sql_query(q, conn, params=[retailer] + week_starts)

def set_units_auto_from_upload(conn, week_start: date, week_end: date, retailer: str, units_by_sku: dict):
    now = datetime.utcnow().isoformat(timespec="seconds") + "Z"
    cur = conn.cursor()
    for sku, units in units_by_sku.items():
        sku = str(sku).strip()
        try:
            units_val = float(units)
        except Exception:
            continue
        cur.execute("""
            INSERT INTO weekly_results(week_start, week_end, retailer, sku,
                                      units_auto, units_override, sales_manual, notes, updated_at)
            VALUES(?,?,?,?,?,?,?,?,?)
            ON CONFLICT(week_start, retailer, sku) DO UPDATE SET
                week_end=excluded.week_end,
                units_auto=excluded.units_auto,
                updated_at=excluded.updated_at
        """, (
            week_start.isoformat(), week_end.isoformat(), retailer, sku,
            units_val, None, None, None, now
        ))
    conn.commit()

# -----------------------------
# Upload parser (based on your example workbook)
# -----------------------------
def parse_weekly_workbook(file, sheet_name: str):
    wb = load_workbook(file, data_only=True)
    if sheet_name not in wb.sheetnames:
        return {}
    ws = wb[sheet_name]
    out = {}

    def add(sku, qty):
        if sku is None:
            return
        sku = str(sku).strip()
        if sku == "" or sku.lower() == "sku":
            return
        try:
            q = float(qty)
        except Exception:
            return
        out[sku] = out.get(sku, 0.0) + q

    if sheet_name in ("Depot", "Lowe's"):
        for r in range(1, ws.max_row + 1):
            add(ws.cell(r, 5).value, ws.cell(r, 6).value)  # E, F

    elif sheet_name == "Depot SO":
        for r in range(1, ws.max_row + 1):
            add(ws.cell(r, 4).value, ws.cell(r, 5).value)  # D, E

    elif sheet_name == "Amazon":
        for r in range(1, ws.max_row + 1):
            add(ws.cell(r, 3).value, ws.cell(r, 14).value)  # C, N

    elif sheet_name == "TSC":
        header_row = None
        for r in range(1, min(ws.max_row, 10) + 1):
            a = ws.cell(r, 1).value
            b = ws.cell(r, 2).value
            if isinstance(a, str) and isinstance(b, str) and "Vendor" in a and "Qty" in b:
                header_row = r
                break
        start = (header_row + 1) if header_row else 2
        for r in range(start, ws.max_row + 1):
            add(ws.cell(r, 1).value, ws.cell(r, 2).value)

    return {k: v for k, v in out.items() if v != 0}

# -----------------------------
# Build multi-week view dataframe
# -----------------------------
def build_multiweek_df(conn, retailer: str, week_meta: list[tuple[date,date,str]], display_labels: list[str], edit_label: str):
    mapping = get_mapping_for_retailer(conn, retailer)
    if mapping.empty:
        return pd.DataFrame()

    label_to_start = {lbl: start.isoformat() for start, _, lbl in week_meta}
    starts = [label_to_start[lbl] for lbl in display_labels if lbl in label_to_start]
    wk = get_week_records(conn, retailer, starts)

    # resolved units per (week_start, sku)
    if not wk.empty:
        wk["UnitsResolved"] = wk["units_override"].where(wk["units_override"].notna(), wk["units_auto"])
    else:
        wk = pd.DataFrame(columns=["week_start","sku","UnitsResolved","sales_manual","notes"])

    base = mapping.rename(columns={"vendor":"Vendor","sku":"SKU","unit_price":"Unit Price"}).copy()
    base["Unit Price"] = pd.to_numeric(base["Unit Price"], errors="coerce")

    # Add per-week columns
    for lbl in display_labels:
        ws = label_to_start.get(lbl)
        if not ws:
            base[lbl] = pd.NA
            continue
        sub = wk[wk["week_start"] == ws][["sku","UnitsResolved"]].rename(columns={"sku":"SKU", "UnitsResolved": lbl})
        base = base.merge(sub, on="SKU", how="left")

    # Add Sales/Notes for edit week only (far right)
    edit_start = label_to_start.get(edit_label)
    if edit_start and not wk.empty:
        sub2 = wk[wk["week_start"] == edit_start][["sku","sales_manual","notes"]].rename(columns={"sku":"SKU"})
        base = base.merge(sub2, on="SKU", how="left")
    else:
        base["sales_manual"] = pd.NA
        base["notes"] = pd.NA

    base.rename(columns={"sales_manual":"Sales", "notes":"Notes"}, inplace=True)

    # Total $ across displayed weeks (calculated, read-only)
    # Sum units across displayed labels * unit price per row
    units_sum = None
    for lbl in display_labels:
        col = pd.to_numeric(base[lbl], errors="coerce")
        units_sum = col if units_sum is None else units_sum.add(col, fill_value=0)
        # Total $ is the Δ $ between the last two displayed weeks: (last - prev) × Unit Price
    if len(display_labels) >= 2:
        prev_lbl = display_labels[-2]
        last_lbl = display_labels[-1]
        prev_u = pd.to_numeric(base[prev_lbl], errors="coerce").fillna(0)
        last_u = pd.to_numeric(base[last_lbl], errors="coerce").fillna(0)
        base["Total $ (Units x Price)"] = ((last_u - prev_u) * base["Unit Price"].fillna(0)).where(base["Unit Price"].notna(), pd.NA)
        base["Total $ (Units x Price)"] = pd.to_numeric(base["Total $ (Units x Price)"], errors="coerce").round(2)
    else:
        base["Total $ (Units x Price)"] = pd.NA
    # Δ Units between the last two displayed weeks (per SKU)
    if len(display_labels) >= 2:
        prev_lbl = display_labels[-2]
        last_lbl = display_labels[-1]
        prev_vals = pd.to_numeric(base[prev_lbl], errors="coerce").fillna(0)
        last_vals = pd.to_numeric(base[last_lbl], errors="coerce").fillna(0)
        base["Δ Units (Last - Prev)"] = last_vals - prev_vals
    else:
        base["Δ Units (Last - Prev)"] = pd.NA



    # Reorder columns: Vendor, SKU, Unit Price, week cols..., Total$, Sales, Notes
    cols = ["Vendor","SKU","Unit Price"] + display_labels + ["Total $ (Units x Price)","Sales","Notes","Δ Units (Last - Prev)"]
    return base[cols]

def save_edit_week(conn, retailer: str, week_start: date, week_end: date, edit_label: str, edited_df: pd.DataFrame):
    now = datetime.utcnow().isoformat(timespec="seconds") + "Z"
    cur = conn.cursor()

    # The editable units are in the column named edit_label.
    for _, row in edited_df.iterrows():
        sku = str(row["SKU"]).strip()
        units_val = row.get(edit_label)
        sales_val = row.get("Sales")
        notes_val = row.get("Notes")

        # Units override stored as the edited value (can be blank to clear)
        units_override = None
        if units_val is not None and not (isinstance(units_val, float) and pd.isna(units_val)):
            try:
                units_override = float(units_val)
            except Exception:
                units_override = None

        sales_manual = None
        if sales_val is not None and not (isinstance(sales_val, float) and pd.isna(sales_val)):
            try:
                sales_manual = float(sales_val)
            except Exception:
                sales_manual = None

        notes_txt = None
        if notes_val is not None and not (isinstance(notes_val, float) and pd.isna(notes_val)) and str(notes_val).strip() != "":
            notes_txt = str(notes_val)

        cur.execute("""
            INSERT INTO weekly_results(week_start, week_end, retailer, sku,
                                      units_auto, units_override, sales_manual, notes, updated_at)
            VALUES(?,?,?,?,?,?,?,?,?)
            ON CONFLICT(week_start, retailer, sku) DO UPDATE SET
                week_end=excluded.week_end,
                units_override=excluded.units_override,
                sales_manual=excluded.sales_manual,
                notes=excluded.notes,
                updated_at=excluded.updated_at
        """, (
            week_start.isoformat(), week_end.isoformat(), retailer, sku,
            None, units_override, sales_manual, notes_txt, now
        ))
    conn.commit()

# -----------------------------
# UI
# -----------------------------
# -----------------------------
# UI
# -----------------------------
st.set_page_config(page_title=APP_TITLE, layout="wide")
st.markdown(BIGGER_TABLES_CSS, unsafe_allow_html=True)
st.markdown(COMPACT_TABLE_CSS, unsafe_allow_html=True)
st.markdown(SPLIT_TABLE_CSS, unsafe_allow_html=True)
st.title(APP_TITLE)


# compact grid
st.markdown(
    """
    <style>
      /* compact grid */
      div[data-testid="stDataFrame"] thead tr th, 
      div[data-testid="stDataFrame"] tbody tr td {
        padding-top: 0.15rem !important;
        padding-bottom: 0.15rem !important;
        padding-left: 0.35rem !important;
        padding-right: 0.35rem !important;
        white-space: nowrap !important;
      }
      /* allow table to size to content */
      div[data-testid="stDataFrame"] [role="grid"] {
        width: max-content !important;
      }
    </style>
    """,
    unsafe_allow_html=True
)


# Light padding so the table fits more rows
st.markdown(
    """
    <style>
      .block-container { padding-top: 1.0rem; padding-bottom: 1.0rem; }
      section[data-testid="stSidebar"] .block-container { padding-top: 0.8rem; }
    </style>
    """,
    unsafe_allow_html=True
)

conn = get_conn()
init_db(conn)
ensure_mapping_loaded(conn, os.path.join(os.path.dirname(__file__), "Vendor-SKU Map.xlsx"))
booted = bootstrap_default_mapping(conn)
refreshed_prices = refresh_mapping_from_bundled_if_needed(conn)

# Load retailers ONCE, before any sidebar/main references
retailers = get_retailers(conn)
if not retailers:
    st.warning("No retailers found yet. You can still use Backup / Restore to load a saved database.")
    retailers = []

week_meta = weeks_2026()
labels = [w[2] for w in week_meta]

# -----------------------------
# Sidebar controls (retailer/weeks/upload)
# -----------------------------

# Ensure `retailer` is defined before sidebar UI-state restore runs (Report tab also has its own dropdown).
try:
    _ret_df = pd.read_sql_query('SELECT DISTINCT retailer FROM weekly_results', conn)
    _ret_list = sorted(_ret_df['retailer'].dropna().unique().tolist())
except Exception:
    _ret_list = []
if _ret_list:
    retailer = st.session_state.get('retailer_report_tab', _ret_list[0])
else:
    retailer = st.session_state.get('retailer_report_tab', '')

with st.sidebar:
    st.header("Setup (optional)")
    st.caption("Vendor map is bundled in the repo. Upload only if you want to replace it.")
    st.write("✅ Loaded bundled mapping" if booted else "ℹ️ Using existing mapping in database")
    if refreshed_prices:
        st.success("Prices refreshed from bundled vendor map")

    map_file = st.file_uploader("Upload Vendor-SKU Map (.xlsx) (optional)", type=["xlsx"])
    if map_file is not None:
        try:
            df_map = pd.read_excel(map_file, sheet_name=0)
            upsert_mapping(conn, df_map)
            st.success("Mapping updated. Reloading retailers…")
            retailers = get_retailers(conn)
        except Exception as e:
            st.error(f"Mapping upload failed: {e}")

    st.divider()
    st.header("Report controls")

    # ---- Year filter (applies to scorecards / totals / saved views) ----
    years = get_available_years(conn)
    year_options = ["All"] + years
    saved_year = get_ui_state(conn, "global::year", default="All")
    if saved_year not in year_options:
        saved_year = "All"
    year_sel = st.selectbox("Year", year_options, index=year_options.index(saved_year), key="year_selector")
    set_ui_state(conn, "global::year", year_sel)




    # Restore saved selections per retailer
    state_key = f"ui::{retailer}"
    saved = get_ui_state(conn, state_key, default={}) or {}
    saved_display = saved.get("display_weeks")
    saved_edit = saved.get("edit_week")

    default_display = labels[:9]  # partial + first 8 full weeks
    if isinstance(saved_display, list):
        default_display = [w for w in labels if w in saved_display] or default_display

    selection_mode = st.radio("Select by", ["Weeks", "Months", "Both"], index=0 if saved.get("selection_mode") is None else ["Weeks","Months","Both"].index(saved.get("selection_mode")), horizontal=True)

    saved_months = saved.get("months") if isinstance(saved.get("months"), list) else []
    month_labels = [MONTH_NAMES[m-1] for m in range(1, 13)]
    month_label_to_num = {MONTH_NAMES[m-1]: m for m in range(1, 13)}
    default_month_labels = [MONTH_NAMES[m-1] for m in saved_months if 1 <= m <= 12]
    month_sel_labels = st.multiselect("Months", month_labels, default=default_month_labels, key="months_sidebar")
    months_selected = [month_label_to_num[x] for x in month_sel_labels]

    display_weeks_manual = st.multiselect("Weeks to display", labels, default=default_display, key="display_weeks_sidebar")
    display_weeks_manual = [lbl for lbl in labels if lbl in display_weeks_manual]  # chronological

    display_weeks_by_month = weeks_for_months(week_meta, months_selected)
    display_weeks_by_month = [lbl for lbl in labels if lbl in display_weeks_by_month]  # chronological

    if selection_mode == "Weeks":
        display_weeks = display_weeks_manual
    elif selection_mode == "Months":
        display_weeks = display_weeks_by_month
    else:
        display_weeks = [lbl for lbl in labels if (lbl in set(display_weeks_manual) or lbl in set(display_weeks_by_month))]

    edit_index = labels.index(saved_edit) if (saved_edit in labels) else 0
    edit_week = st.selectbox("Week to edit", labels, index=edit_index, key="edit_week_sidebar")

    if edit_week not in display_weeks:
        display_weeks = display_weeks + [edit_week]

    # Persist selections every run
    set_ui_state(conn, state_key, {"display_weeks": display_weeks, "edit_week": edit_week, "months": months_selected, "selection_mode": selection_mode})

    st.divider()
    st.subheader("Upload units (APP workbook)")
    app_file = st.file_uploader("Weekly APP workbook (.xlsx)", type=["xlsx"], key="app_units_upload")

    if app_file is not None:
        parsed_label, _, _ = parse_week_label_from_filename(getattr(app_file, "name", ""))
        if parsed_label:
            st.caption(f"Detected week in filename: {parsed_label}")

        if st.button("Import units into Edit Week", type="primary", use_container_width=True):
            chosen_start, chosen_end, _ = next((a, b, l) for a, b, l in week_meta if l == edit_week)

            wb_up = load_workbook(app_file, data_only=True)
            imported = []
            skipped = []
            for sh in wb_up.sheetnames:
                retailer_name = normalize_retailer_sheet_name(sh)
                ws = wb_up[sh]
                units = parse_app_aggregated_sheet(ws)
                if not units:
                    skipped.append(sh)
                    continue
                set_units_auto_from_upload(conn, chosen_start, chosen_end, retailer_name, units)
                imported.append((retailer_name, len(units)))

            if imported:
                msg = ", ".join([f"{r} ({n})" for r, n in imported])
                st.success(f"Imported: {msg}")
            if skipped:
                st.caption(f"Skipped empty sheets: {', '.join(skipped)}")


# Global year filter selection
year_filter = st.session_state.get('year_selector') or get_ui_state(conn, 'global::year', default='All') or 'All'

# -----------------------------
# Main tabs
# -----------------------------

# Global toggles (apply across all tabs)
if "global_edit_mode" not in st.session_state:
    st.session_state["global_edit_mode"] = True
if "global_color_deltas" not in st.session_state:
    st.session_state["global_color_deltas"] = True

st.toggle("Edit mode (applies to all tabs)", key="global_edit_mode")
st.toggle("Color positive/negative deltas", key="global_color_deltas")

(tab_summary,
 tab_reports,
 tab_saved_retailer,
 tab_saved_vendor,
 tab_retailer_totals,
 tab_retailer_scorecard,
 tab_retailer_scorecard_summary,
 tab_vendor_totals,
 tab_vendor_scorecard,
 tab_vendor_scorecard_summary,
 tab_sku_lifecycle,
 tab_health,
 tab_backup) = st.tabs([
    "Summary",
    "Reports",
    "Saved Views – Retailers",
    "Saved Views – Vendors",
    "Retailer Totals",
    "Retailer Scorecard",
    "Retailer Scorecard Summary",
    "Vendor Totals",
    "Vendor Scorecard",
    "Vendor Scorecard Summary",
    "SKU Lifecycle",
    "Health Flags",
    "Backup / Restore",
])


with tab_reports:
    # Retailer selector (inside the Report tab)
    retailers = sorted(pd.read_sql_query('SELECT DISTINCT retailer FROM weekly_results', conn)['retailer'].dropna().unique().tolist())
    if not retailers:
        st.info('Upload a week workbook first so retailers appear here.')
    else:
        retailer = st.selectbox('Retailer', retailers, key='retailer_report_tab')

    edit_mode = st.session_state.get('global_edit_mode', True)
    color_deltas = st.session_state.get('global_color_deltas', True)
    st.markdown(f"**Retailer:** {retailer}  |  **Edit week:** {edit_week}  |  **Weeks shown:** {', '.join(display_weeks)}")

    # Build and render table
    df = build_multiweek_df(conn, retailer, week_meta, display_weeks, edit_week)

    # Optional filter: only show items with activity
    show_only_with_units = st.checkbox("Show only items with units (or sales)", value=True)
    if show_only_with_units and not df.empty:
        week_cols = [c for c in display_weeks if c in df.columns]
        units_any = (df[week_cols].apply(pd.to_numeric, errors="coerce").fillna(0) > 0).any(axis=1) if week_cols else pd.Series([False] * len(df))
        sales_any = pd.to_numeric(df["Sales"], errors="coerce").fillna(0) != 0
        df = df[units_any | sales_any].reset_index(drop=True)

    if df.empty:
        st.info("No rows for this retailer in your mapping (or no activity for the selected weeks).")
        skip_report = True
    else:
        skip_report = False
    if not skip_report:
    
        # Keep Unit Price for calculations, but hide it from the table
        unit_price = pd.to_numeric(df["Unit Price"], errors="coerce").fillna(0)
        df = df.drop(columns=["Unit Price"])
        # Keep Sales and Notes for persistence, but hide from the table
        hidden_sales = pd.to_numeric(df["Sales"], errors="coerce") if "Sales" in df.columns else None
        df = df.drop(columns=[c for c in ["Sales", "Notes"] if c in df.columns])
    
        # Ensure all money columns are numeric + rounded (prevents long float tails)
        money_cols_all = [c for c in df.columns if "$" in c]
        for c in money_cols_all:
            df[c] = pd.to_numeric(df[c], errors="coerce").round(2)
    
        # Disable columns: keep Vendor/SKU and non-edit weeks read-only
        disabled_cols = ["Vendor", "SKU", "Total $ (Units x Price)", "Δ Units (Last - Prev)"] + [w for w in display_weeks if w != edit_week]
    
        if not edit_mode:
            view_df = df.copy()
    
            # Units table (numeric)
            view_units = view_df.copy()
            # Units table should not include any $ columns (sales are shown in the separate dollars table)
            view_units = view_units[[c for c in view_units.columns if '$' not in c]]
    
            # Dollars table: Vendor, SKU + $ per selected week (units * unit_price)
            view_dollars = pd.DataFrame({
                "Vendor": view_units.get("Vendor", ""),
                "SKU": view_units.get("SKU", ""),
            })
            for w in display_weeks:
                if w in view_units.columns:
                    u = pd.to_numeric(view_units[w], errors="coerce").fillna(0)
                    view_dollars[w] = (u * unit_price).round(2)
    
            # Δ $ between last two weeks
            if len(display_weeks) >= 2:
                prev_w, last_w = display_weeks[-2], display_weeks[-1]
                if prev_w in view_dollars.columns and last_w in view_dollars.columns:
                    view_dollars["Δ $ (Last - Prev)"] = (
                        pd.to_numeric(view_dollars[last_w], errors="coerce").fillna(0)
                        - pd.to_numeric(view_dollars[prev_w], errors="coerce").fillna(0)
                    ).round(2)
                else:
                    view_dollars["Δ $ (Last - Prev)"] = pd.NA
            else:
                view_dollars["Δ $ (Last - Prev)"] = pd.NA
    
            # Currency formatting as strings (isolated to dollars table)
            # Append TOTAL row to both tables (based on shown rows)
            week_cols_units = [w for w in display_weeks if w in view_units.columns]
            tot_units = {w: float(pd.to_numeric(view_units[w], errors='coerce').fillna(0).sum()) for w in week_cols_units}
            if len(week_cols_units) >= 2:
                prev_w, last_w = week_cols_units[-2], week_cols_units[-1]
                delta_units_total = tot_units[last_w] - tot_units[prev_w]
            else:
                delta_units_total = pd.NA
            totals_units_row = {'Vendor': 'TOTAL', 'SKU': ''}
            totals_units_row.update(tot_units)
            if 'Δ Units (Last - Prev)' in view_units.columns:
                totals_units_row['Δ Units (Last - Prev)'] = delta_units_total
            view_units = pd.concat([view_units, pd.DataFrame([totals_units_row])], ignore_index=True)
            
            week_cols_dollars = [w for w in display_weeks if w in view_dollars.columns]
            tot_dollars = {w: float(pd.to_numeric(view_dollars[w], errors='coerce').fillna(0).sum()) for w in week_cols_dollars}
            if len(week_cols_dollars) >= 2:
                prev_w, last_w = week_cols_dollars[-2], week_cols_dollars[-1]
                delta_dollars_total = tot_dollars[last_w] - tot_dollars[prev_w]
            else:
                delta_dollars_total = pd.NA
            totals_dollars_row = {'Vendor': 'TOTAL', 'SKU': ''}
            totals_dollars_row.update(tot_dollars)
            totals_dollars_row['Δ $ (Last - Prev)'] = delta_dollars_total
            view_dollars = pd.concat([view_dollars, pd.DataFrame([totals_dollars_row])], ignore_index=True)
            
            for c in [w for w in display_weeks if w in view_dollars.columns] + ["Δ $ (Last - Prev)"]:
                if c in view_dollars.columns:
                    view_dollars[c] = pd.to_numeric(view_dollars[c], errors="coerce").round(2).apply(fmt_currency_str)
    
            def _color_pos_neg(val):
                s = str(val)
                try:
                    neg = s.strip().startswith('(') and s.strip().endswith(')')
                    s2 = s.replace('(', '').replace(')', '').replace('$', '').replace(',', '')
                    v = float(s2) if s2.strip() != '' else 0.0
                    if neg:
                        v = -abs(v)
                except Exception:
                    return ''
                if v > 0:
                    return 'color: #1f8b4c; font-weight: 600;'
                if v < 0:
                    return 'color: #c92a2a; font-weight: 600;'
                return ''
    
            # Render two tables side-by-side
            left_col, right_col = st.columns([1, 1], gap="small")
    
            with left_col:
                styled_units = view_units.style
                if color_deltas and "Δ Units (Last - Prev)" in view_units.columns:
                    styled_units = styled_units.applymap(_color_pos_neg, subset=["Δ Units (Last - Prev)"])
                st.dataframe(
                    styled_units,
                    use_container_width=True,
                    height=1200,
                    column_config={
                        "Vendor": st.column_config.TextColumn(width="small"),
                        "SKU": st.column_config.TextColumn(width="small"),
                        **{w: st.column_config.NumberColumn(format="%.0f", width="small") for w in display_weeks if w in view_units.columns},
                        "Δ Units (Last - Prev)": st.column_config.NumberColumn(format="%.0f", width="small"),
                    },
                )
    
            with right_col:
                styled_dollars = view_dollars.style
                if color_deltas and "Δ $ (Last - Prev)" in view_dollars.columns:
                    styled_dollars = styled_dollars.applymap(_color_pos_neg, subset=["Δ $ (Last - Prev)"])
                st.dataframe(
                    styled_dollars,
                    use_container_width=True,
                    height=1200,
                    hide_index=True,
                    column_config={
                        "Vendor": st.column_config.TextColumn(width="small"),
                        "SKU": st.column_config.TextColumn(width="small"),
                        **{w: st.column_config.TextColumn(width="small") for w in display_weeks if w in view_dollars.columns},
                        "Δ $ (Last - Prev)": st.column_config.TextColumn(width="small"),
                    },
                )
    
            edited = view_df
    
        else:
            # Editor: Units editable (only Edit Week). Dollars computed from the edited units.
            left_col, right_col = st.columns([1, 1], gap="small")
    
            with left_col:
                df_editor_main = df.copy()
                keep_cols = ["Vendor", "SKU"] + [w for w in display_weeks if w in df_editor_main.columns]
                if "Δ Units (Last - Prev)" in df_editor_main.columns:
                    keep_cols += ["Δ Units (Last - Prev)"]
                df_editor_main = df_editor_main[keep_cols].copy()
    
                edited = st.data_editor(
                    df_editor_main,
                    height=1200,
                    use_container_width=True,
                    hide_index=True,
                    disabled=disabled_cols,
                    column_config={
                        "Vendor": st.column_config.TextColumn(width="small"),
                        "SKU": st.column_config.TextColumn(width="small"),
                        **{w: st.column_config.NumberColumn(format="%.0f", width="small") for w in display_weeks if w in df_editor_main.columns},
                        "Δ Units (Last - Prev)": st.column_config.NumberColumn(format="%.0f", width="small"),
                    },
                )
    
            with right_col:
                dollars = pd.DataFrame({
                    "Vendor": edited.get("Vendor", ""),
                    "SKU": edited.get("SKU", ""),
                })
                for w in display_weeks:
                    if w in edited.columns:
                        u = pd.to_numeric(edited[w], errors="coerce").fillna(0)
                        dollars[w] = (u * unit_price).round(2)
    
                if len(display_weeks) >= 2:
                    prev_w, last_w = display_weeks[-2], display_weeks[-1]
                    if prev_w in dollars.columns and last_w in dollars.columns:
                        dollars["Δ $ (Last - Prev)"] = (
                            pd.to_numeric(dollars[last_w], errors="coerce").fillna(0)
                            - pd.to_numeric(dollars[prev_w], errors="coerce").fillna(0)
                        ).round(2)
                    else:
                        dollars["Δ $ (Last - Prev)"] = pd.NA
                else:
                    dollars["Δ $ (Last - Prev)"] = pd.NA
    
                # Append TOTAL row (dollars table) based on edited rows
                week_cols_dollars = [w for w in display_weeks if w in dollars.columns]
                tot_dollars = {w: float(pd.to_numeric(dollars[w], errors='coerce').fillna(0).sum()) for w in week_cols_dollars}
                if len(week_cols_dollars) >= 2:
                    prev_w, last_w = week_cols_dollars[-2], week_cols_dollars[-1]
                    delta_dollars_total = tot_dollars[last_w] - tot_dollars[prev_w]
                else:
                    delta_dollars_total = pd.NA
                totals_row = {'Vendor': 'TOTAL', 'SKU': ''}
                totals_row.update(tot_dollars)
                totals_row['Δ $ (Last - Prev)'] = delta_dollars_total
                dollars = pd.concat([dollars, pd.DataFrame([totals_row])], ignore_index=True)
                
                for c in [w for w in display_weeks if w in dollars.columns] + ["Δ $ (Last - Prev)"]:
                    if c in dollars.columns:
                        dollars[c] = pd.to_numeric(dollars[c], errors="coerce").round(2).apply(fmt_currency_str)
    
                st.dataframe(
                    dollars,
                    use_container_width=True,
                    height=1200,
                    hide_index=True,
                    column_config={
                        "Vendor": st.column_config.TextColumn(width="small"),
                        "SKU": st.column_config.TextColumn(width="small"),
                        **{w: st.column_config.TextColumn(width="small") for w in display_weeks if w in dollars.columns},
                        "Δ $ (Last - Prev)": st.column_config.TextColumn(width="small"),
                    },
                )
    
        c1, c2 = st.columns([1, 3])
    
        with c1:
            if st.button("Save edits", type="primary"):
                start, end, _ = next((a, b, l) for a, b, l in week_meta if l == edit_week)
                save_edit_week(conn, retailer, start, end, edit_week, edited)
                st.success("Saved.")
        with c2:
            st.caption("Only the selected Edit Week column is editable. Far-right column shows Δ Units (last selected week minus the previous week).")
    
    
with tab_summary:
    st.subheader("Summary")
    st.caption("Totals per retailer by week (Sales $). Includes Δ $ (Last - Prev) and a Total row at the bottom.")

    selected_labels = display_weeks
    label_to_start = {lbl: start.isoformat() for start, _, lbl in week_meta}
    selected_starts = [label_to_start[lbl] for lbl in selected_labels if lbl in label_to_start]

    if not selected_starts:
        st.info("Select at least one week.")
    else:
        placeholders = ",".join(["?"] * len(selected_starts))
        wk = pd.read_sql_query(
            f'''
            SELECT week_start, retailer, sku, units_auto, units_override
            FROM weekly_results
            WHERE week_start IN ({placeholders})
            ''',
            conn,
            params=selected_starts
        )

        if wk.empty:
            st.info("No data found for the selected weeks yet.")
        else:
            wk["Units"] = wk["units_override"].where(wk["units_override"].notna(), wk["units_auto"])
            wk["Units"] = pd.to_numeric(wk["Units"], errors="coerce").fillna(0)

            mapping_all = pd.read_sql_query(
                '''
                SELECT retailer, sku, unit_price
                FROM sku_mapping
                WHERE active = 1
                ''',
                conn
            )
            mapping_all["unit_price"] = pd.to_numeric(mapping_all["unit_price"], errors="coerce").fillna(0)

            dfm = wk.merge(mapping_all, on=["retailer", "sku"], how="left")
            dfm["unit_price"] = pd.to_numeric(dfm["unit_price"], errors="coerce").fillna(0)
            dfm["Sales"] = (dfm["Units"] * dfm["unit_price"]).round(2)

            # totals by retailer + week_start
            agg = dfm.groupby(["retailer", "week_start"], as_index=False).agg(
                Sales=("Sales", "sum"),
            )

            # map week_start back to label for display
            start_to_label = {start.isoformat(): lbl for start, _, lbl in week_meta}
            agg["Week"] = agg["week_start"].map(start_to_label)

            pivot = agg.pivot_table(index="retailer", columns="Week", values="Sales", aggfunc="sum", fill_value=0)

            # ensure column order matches selected_labels
            ordered = [lbl for lbl in selected_labels if lbl in pivot.columns]
            pivot = pivot[ordered].copy()

            # add delta column (last - prev)
            if len(ordered) >= 2:
                prev_lbl, last_lbl = ordered[-2], ordered[-1]
                pivot["Δ $ (Last - Prev)"] = (pivot[last_lbl] - pivot[prev_lbl]).round(2)
            else:
                pivot["Δ $ (Last - Prev)"] = 0.0

            pivot = pivot.reset_index().rename(columns={"retailer": "Retailer"})

            # Add total row at bottom
            total_row = {"Retailer": "Total"}
            for lbl in ordered:
                total_row[lbl] = float(pivot[lbl].sum())
            total_row["Δ $ (Last - Prev)"] = float(pivot["Δ $ (Last - Prev)"].sum())
            pivot = pd.concat([pivot, pd.DataFrame([total_row])], ignore_index=True)

            # format currency columns
            for lbl in ordered + ["Δ $ (Last - Prev)"]:
                pivot[lbl] = pd.to_numeric(pivot[lbl], errors="coerce").round(2).apply(fmt_currency_str)

            # Optional coloring for delta
            def _color_pos_neg(val):
                s = str(val)
                try:
                    neg = s.strip().startswith('(') and s.strip().endswith(')')
                    s2 = s.replace('(', '').replace(')', '').replace('$', '').replace(',', '')
                    v = float(s2) if s2.strip() != '' else 0.0
                    if neg:
                        v = -abs(v)
                except Exception:
                    return ""
                if v > 0:
                    return "color: #1f8b4c; font-weight: 600;"
                if v < 0:
                    return "color: #c92a2a; font-weight: 600;"
                return ""

            styled = pivot.style
            if "Δ $ (Last - Prev)" in pivot.columns:
                styled = styled.applymap(_color_pos_neg, subset=["Δ $ (Last - Prev)"])

            st.dataframe(
                styled,
                use_container_width=True,
                height=650,
                hide_index=True,
                column_config={c: st.column_config.TextColumn(width="small") for c in pivot.columns},
            )
with tab_saved_retailer:
    st.subheader("Saved Views — Retailers")
    st.caption("Dynamic ranges: pick a retailer and show Last 2 / 4 / 8 weeks. This does NOT change your manual selections on the Report tab.")

    wk_avail = pd.read_sql_query(
        "SELECT DISTINCT week_start FROM weekly_results ORDER BY week_start",
        conn
    )
    week_starts = wk_avail["week_start"].dropna().astype(str).tolist()
    week_starts = [ws for ws in week_starts if (year_filter == "All" or str(ws).startswith(str(year_filter)))]

    start_to_label = {start.isoformat(): lbl for start, _, lbl in week_meta}

    retailers = sorted(pd.read_sql_query("SELECT DISTINCT retailer FROM weekly_results", conn)["retailer"].dropna().unique().tolist())
    if not retailers or not week_starts:
        st.info("Upload at least one week workbook first so retailers and weeks are available.")
    else:
        r = st.selectbox("Retailer", retailers, key="saved_view_retailer")
        range_opt = st.selectbox("Range", ["Last 2 Weeks", "Last 4 Weeks", "Last 8 Weeks"], key="saved_view_retailer_range")
        n = 2 if "2" in range_opt else (4 if "4" in range_opt else 8)

        sel_starts = week_starts[-n:]
        sel_labels = [start_to_label.get(ws, ws) for ws in sel_starts]

        placeholders = ",".join(["?"] * len(sel_starts))
        wk = pd.read_sql_query(
            f'''
            SELECT week_start, retailer, sku, units_auto, units_override
            FROM weekly_results
            WHERE retailer = ? AND week_start IN ({placeholders})
            ''',
            conn,
            params=[r] + sel_starts
        )

        mapping = pd.read_sql_query(
            "SELECT retailer, vendor, sku, unit_price FROM sku_mapping WHERE active = 1 AND retailer = ?",
            conn,
            params=(r,)
        )
        mapping["unit_price"] = pd.to_numeric(mapping["unit_price"], errors="coerce").fillna(0)

        if wk.empty:
            st.info("No data for that retailer in the selected range.")
        else:
            wk["Units"] = wk["units_override"].where(wk["units_override"].notna(), wk["units_auto"])
            wk["Units"] = pd.to_numeric(wk["Units"], errors="coerce").fillna(0)

            dfm = wk.merge(mapping, on=["retailer","sku"], how="left")
            dfm["vendor"] = dfm["vendor"].fillna("Unknown")
            dfm["unit_price"] = pd.to_numeric(dfm["unit_price"], errors="coerce").fillna(0)
            dfm["Sales"] = (dfm["Units"] * dfm["unit_price"]).round(2)

            dfm["Week"] = dfm["week_start"].map(lambda ws: start_to_label.get(str(ws), str(ws)))

            key_cols = ["vendor","sku"]
            units_p = dfm.pivot_table(index=key_cols, columns="Week", values="Units", aggfunc="sum", fill_value=0)
            sales_p = dfm.pivot_table(index=key_cols, columns="Week", values="Sales", aggfunc="sum", fill_value=0)

            units_p = units_p[[c for c in sel_labels if c in units_p.columns]].copy()
            sales_p = sales_p[[c for c in sel_labels if c in sales_p.columns]].copy()

            if len(sel_labels) >= 2:
                prev_lbl, last_lbl = sel_labels[-2], sel_labels[-1]
                units_p["Δ Units (Last - Prev)"] = units_p[last_lbl] - units_p[prev_lbl]
                sales_p["Δ $ (Last - Prev)"] = (sales_p[last_lbl] - sales_p[prev_lbl]).round(2)
            else:
                units_p["Δ Units (Last - Prev)"] = 0
                sales_p["Δ $ (Last - Prev)"] = 0.0

            units_df = units_p.reset_index().rename(columns={"vendor":"Vendor","sku":"SKU"})
            sales_df = sales_p.reset_index().rename(columns={"vendor":"Vendor","sku":"SKU"})

            tot_units = {"Vendor":"Total","SKU":""}
            for c in sel_labels + ["Δ Units (Last - Prev)"]:
                tot_units[c] = float(units_df[c].sum()) if c in units_df.columns else 0
            units_df = pd.concat([units_df, pd.DataFrame([tot_units])], ignore_index=True)

            tot_sales = {"Vendor":"Total","SKU":""}
            for c in sel_labels + ["Δ $ (Last - Prev)"]:
                tot_sales[c] = float(pd.to_numeric(sales_df[c], errors="coerce").sum()) if c in sales_df.columns else 0.0
            sales_df = pd.concat([sales_df, pd.DataFrame([tot_sales])], ignore_index=True)

            for c in sel_labels + ["Δ $ (Last - Prev)"]:
                if c in sales_df.columns:
                    sales_df[c] = pd.to_numeric(sales_df[c], errors="coerce").round(2).apply(fmt_currency_str)

            def _color_pos_neg(val):
                s = str(val)
                try:
                    neg = s.strip().startswith("(") and s.strip().endswith(")")
                    s2 = s.replace("(", "").replace(")", "").replace("$", "").replace(",", "")
                    v = float(s2) if s2.strip() else 0.0
                    if neg:
                        v = -abs(v)
                except Exception:
                    return ""
                if v > 0:
                    return "color: #1f8b4c; font-weight: 600;"
                if v < 0:
                    return "color: #c92a2a; font-weight: 600;"
                return ""

            left_col, right_col = st.columns([1, 1], gap="small")
            with left_col:
                st.markdown("#### Units")
                styled_u = units_df.style.applymap(_color_pos_neg, subset=["Δ Units (Last - Prev)"])
                st.dataframe(
                    styled_u,
                    use_container_width=True,
                    height=750,
                    hide_index=True,
                    column_config={
                        "Vendor": st.column_config.TextColumn(width="small"),
                        "SKU": st.column_config.TextColumn(width="small"),
                        **{c: st.column_config.NumberColumn(format="%.0f", width="small") for c in sel_labels if c in units_df.columns},
                        "Δ Units (Last - Prev)": st.column_config.NumberColumn(format="%.0f", width="small"),
                    },
                )

            with right_col:
                st.markdown("#### Sales ($)")
                styled_s = sales_df.style.applymap(_color_pos_neg, subset=["Δ $ (Last - Prev)"])
                st.dataframe(
                    styled_s,
                    use_container_width=True,
                    height=750,
                    hide_index=True,
                    column_config={
                        "Vendor": st.column_config.TextColumn(width="small"),
                        "SKU": st.column_config.TextColumn(width="small"),
                        **{c: st.column_config.TextColumn(width="small") for c in sel_labels if c in sales_df.columns},
                        "Δ $ (Last - Prev)": st.column_config.TextColumn(width="small"),
                    },
                )

with tab_saved_vendor:
    st.subheader("Saved Views — Vendors")
    st.caption("Dynamic ranges: pick a vendor and show its Top 10 SKUs across all retailers for Last 2 / 4 / 8 weeks. This does NOT change your manual selections on the other tabs.")

    wk_avail = pd.read_sql_query(
        "SELECT DISTINCT week_start FROM weekly_results ORDER BY week_start",
        conn
    )
    week_starts = wk_avail["week_start"].dropna().astype(str).tolist()
    week_starts = [ws for ws in week_starts if (year_filter == "All" or str(ws).startswith(str(year_filter)))]

    vendors = sorted(pd.read_sql_query("SELECT DISTINCT vendor FROM sku_mapping WHERE active = 1", conn)["vendor"].dropna().unique().tolist())
    if not vendors or not week_starts:
        st.info("Upload at least one week workbook first so vendors and weeks are available.")
    else:
        vsel = st.selectbox("Vendor", vendors, key="saved_view_vendor")
        range_opt = st.selectbox("Range", ["Last 2 Weeks", "Last 4 Weeks", "Last 8 Weeks"], key="saved_view_vendor_range")
        n = 2 if "2" in range_opt else (4 if "4" in range_opt else 8)
        sel_starts = week_starts[-n:]

        placeholders = ",".join(["?"] * len(sel_starts))
        wk = pd.read_sql_query(
            f'''
            SELECT week_start, retailer, sku, units_auto, units_override
            FROM weekly_results
            WHERE week_start IN ({placeholders})
            ''',
            conn,
            params=sel_starts
        )

        mapping_all = pd.read_sql_query(
            "SELECT retailer, vendor, sku, unit_price FROM sku_mapping WHERE active = 1 AND vendor = ?",
            conn,
            params=(vsel,)
        )
        mapping_all["unit_price"] = pd.to_numeric(mapping_all["unit_price"], errors="coerce").fillna(0)

        if wk.empty or mapping_all.empty:
            st.info("No data for that vendor in the selected range.")
        else:
            wk["Units"] = wk["units_override"].where(wk["units_override"].notna(), wk["units_auto"])
            wk["Units"] = pd.to_numeric(wk["Units"], errors="coerce").fillna(0)

            dfm = wk.merge(mapping_all, on=["retailer","sku"], how="inner")
            dfm["unit_price"] = pd.to_numeric(dfm["unit_price"], errors="coerce").fillna(0)
            dfm["Sales"] = (dfm["Units"] * dfm["unit_price"]).round(2)

            agg = dfm.groupby(["retailer","sku"], as_index=False).agg(
                Units=("Units", "sum"),
                Sales=("Sales", "sum"),
            )

            top = agg.sort_values("Units", ascending=False).head(10)

            left_col, right_col = st.columns([1, 1], gap="small")
            with left_col:
                st.markdown("#### Top 10 — Units")
                t = top[["retailer","sku","Units"]].rename(columns={"retailer":"Retailer","sku":"SKU"})
                st.dataframe(
                    t,
                    use_container_width=True,
                    height=700,
                    hide_index=True,
                    column_config={
                        "Retailer": st.column_config.TextColumn(width="small"),
                        "SKU": st.column_config.TextColumn(width="small"),
                        "Units Rank": st.column_config.NumberColumn(format="%.0f", width="small"),
                        "Units": st.column_config.NumberColumn(format="%.0f", width="small"),
                    },
                )

            with right_col:
                st.markdown("#### Top 10 — Sales ($)")
                t = top[["retailer","sku","Sales"]].rename(columns={"retailer":"Retailer","sku":"SKU"})
                t["Sales"] = pd.to_numeric(t["Sales"], errors="coerce").round(2).apply(fmt_currency_str)
                st.dataframe(
                    t,
                    use_container_width=True,
                    height=700,
                    hide_index=True,
                    column_config={
                        "Retailer": st.column_config.TextColumn(width="small"),
                        "SKU": st.column_config.TextColumn(width="small"),
                        "Sales": st.column_config.TextColumn(width="small"),
                    },
                )



with tab_retailer_totals:
    st.subheader("Retailer Totals (Year)")
    st.caption("Totals across ALL stored weeks for the year. Left = Units, Right = Sales ($).")

    # Pull ALL stored weeks (year-long)
    wk = pd.read_sql_query(
        '''
        SELECT week_start, retailer, sku, units_auto, units_override
        FROM weekly_results
        ''',
        conn
    )

    if wk.empty:
        st.info("No unit data stored yet. Upload at least one weekly workbook.")
    else:
        wk["Units"] = wk["units_override"].where(wk["units_override"].notna(), wk["units_auto"])
        wk["Units"] = pd.to_numeric(wk["Units"], errors="coerce").fillna(0)

        mapping_all = pd.read_sql_query(
            '''
            SELECT retailer, vendor, sku, unit_price
            FROM sku_mapping
            WHERE active = 1
            ''',
            conn
        )
        mapping_all["unit_price"] = pd.to_numeric(mapping_all["unit_price"], errors="coerce").fillna(0)

        dfm = wk.merge(mapping_all, on=["retailer", "sku"], how="left")
        dfm["unit_price"] = pd.to_numeric(dfm["unit_price"], errors="coerce").fillna(0)
        dfm["Sales"] = (dfm["Units"] * dfm["unit_price"]).round(2)

        agg = dfm.groupby("retailer", as_index=False).agg(
            Units=("Units", "sum"),
            Sales=("Sales", "sum"),
        )

        # Rankings (descending)
        agg["Units Rank"] = agg["Units"].rank(method="dense", ascending=False).astype(int)
        agg["Sales Rank"] = agg["Sales"].rank(method="dense", ascending=False).astype(int)

        agg_units = agg[["retailer","Units Rank","Units"]].sort_values("Units", ascending=False).rename(columns={"retailer":"Retailer"})
        agg_sales = agg[["retailer","Sales Rank","Sales"]].sort_values("Sales", ascending=False).rename(columns={"retailer":"Retailer"}).copy()
        agg_sales["Sales"] = pd.to_numeric(agg_sales["Sales"], errors="coerce").round(2).apply(fmt_currency_str)

        # Total rows
        agg_units = pd.concat([agg_units, pd.DataFrame([{"Retailer":"Total","Units": float(agg_units["Units"].sum())}])], ignore_index=True)
        total_sales = float(pd.to_numeric(agg["Sales"], errors="coerce").sum())
        agg_sales = pd.concat([agg_sales, pd.DataFrame([{"Retailer":"Total","Sales": fmt_currency_str(total_sales)}])], ignore_index=True)

        left_col, right_col = st.columns([1, 1], gap="small")
        with left_col:
            st.markdown("#### Total Units by Retailer")
            st.dataframe(
                agg_units,
                use_container_width=True,
                height=750,
                hide_index=True,
                column_config={
                    "Retailer": st.column_config.TextColumn(width="small"),
                    "Units Rank": st.column_config.NumberColumn(format="%.0f", width="small"),
                    "Units": st.column_config.NumberColumn(format="%.0f", width="small"),
                },
            )

        with right_col:
            st.markdown("#### Total Sales ($) by Retailer")
            st.dataframe(
                agg_sales,
                use_container_width=True,
                height=750,
                hide_index=True,
                column_config={
                    "Retailer": st.column_config.TextColumn(width="small"),
                    "Sales Rank": st.column_config.NumberColumn(format="%.0f", width="small"),
                    "Sales": st.column_config.TextColumn(width="small"),
                },
            )



with tab_retailer_scorecard:
    st.subheader("Retailer Scorecard")
    st.caption("Single-retailer view: YTD totals plus WoW and MoM deltas, plus Top SKUs and Vendor breakdown. (No charts)")

    if not table_exists(conn, "weekly_results"):
        st.info("No data loaded yet.")
    else:
        retailers = pd.read_sql_query("SELECT DISTINCT retailer FROM weekly_results", conn)["retailer"]
        retailers = sorted(retailers.dropna().astype(str).str.strip().unique().tolist())

        if not retailers:
            st.info("No retailers found in weekly_results yet.")
        else:
            r_sel = st.selectbox("Retailer", retailers, key="retailer_scorecard_sel")

            wk = pd.read_sql_query(
                "SELECT week_start, retailer, sku, units_auto, units_override FROM weekly_results WHERE retailer = ?",
                conn,
                params=[r_sel],
            )

            wk["week_start"] = pd.to_datetime(wk["week_start"], errors="coerce")
            wk = filter_df_year(wk, year_filter, "week_start")
            wk = wk.dropna(subset=["week_start"])

            if wk.empty:
                st.info("No rows for this retailer in the selected year.")
            else:
                wk["sku"] = wk["sku"].astype(str).str.strip()

                wk["Units"] = wk["units_override"].where(wk["units_override"].notna(), wk["units_auto"])
                wk["Units"] = pd.to_numeric(wk["Units"], errors="coerce").fillna(0)

                # Join price/vendor from sku_mapping if available (dedupe per SKU for this retailer)
                vendor_col = "vendor"
                price_col = "unit_price"
                if table_exists(conn, "sku_mapping"):
                    try:
                        m = pd.read_sql_query("SELECT retailer, vendor, sku, unit_price FROM sku_mapping", conn)
                        m["retailer"] = m["retailer"].astype(str).str.strip()
                        m["sku"] = m["sku"].astype(str).str.strip()
                        m["vendor"] = m["vendor"].astype(str).str.strip()
                        m["unit_price"] = pd.to_numeric(m["unit_price"], errors="coerce")
                        m = m[m["retailer"] == r_sel].copy()
                        if not m.empty:
                            # keep first non-null per sku
                            m = (m.sort_values(["sku"])
                                   .groupby("sku", as_index=False)
                                   .agg(vendor=("vendor","first"), unit_price=("unit_price","first")))
                            wk = wk.merge(m, on="sku", how="left")
                    except Exception:
                        pass

                wk["vendor"] = wk.get("vendor", pd.Series(["Unknown"] * len(wk))).fillna("Unknown")
                wk["unit_price"] = pd.to_numeric(wk.get("unit_price", 0.0), errors="coerce").fillna(0.0)
                wk["Sales"] = (wk["Units"] * wk["unit_price"]).fillna(0.0)

                # KPIs
                ytd_units = float(wk["Units"].sum())
                ytd_sales = float(wk["Sales"].sum())

                wk["week_str"] = wk["week_start"].dt.strftime("%Y-%m-%d")
                weeks = sorted(wk["week_str"].dropna().unique().tolist())
                if len(weeks) >= 2:
                    prev_w, last_w = weeks[-2], weeks[-1]
                    prev = wk[wk["week_str"] == prev_w][["Units","Sales"]].sum(numeric_only=True)
                    last = wk[wk["week_str"] == last_w][["Units","Sales"]].sum(numeric_only=True)
                    wow_units = float(last.get("Units", 0) - prev.get("Units", 0))
                    wow_sales = float(last.get("Sales", 0) - prev.get("Sales", 0))
                else:
                    wow_units = 0.0
                    wow_sales = 0.0

                cur_month = wk["week_start"].max().to_period("M")
                prev_month = cur_month - 1
                m_cur = wk[wk["week_start"].dt.to_period("M") == cur_month][["Units","Sales"]].sum(numeric_only=True)
                m_prev = wk[wk["week_start"].dt.to_period("M") == prev_month][["Units","Sales"]].sum(numeric_only=True)
                mom_units = float(m_cur.get("Units", 0) - m_prev.get("Units", 0))
                mom_sales = float(m_cur.get("Sales", 0) - m_prev.get("Sales", 0))

                c1, c2, c3, c4 = st.columns(4, gap="small")
                c1.metric("Units YTD", f"{ytd_units:,.0f}", f"{wow_units:+,.0f} WoW")
                c2.metric("Sales YTD", f"${ytd_sales:,.2f}", f"${wow_sales:+,.2f} WoW")
                c3.metric("Units MoM Δ", f"{mom_units:,.0f}")
                c4.metric("Sales MoM Δ", f"${mom_sales:,.2f}")

                # Weekly trend (table only)
                st.markdown("### Weekly trend")
                trend = (wk.groupby("week_str", as_index=False)
                           .agg(Units=("Units","sum"), Sales=("Sales","sum"))
                           .sort_values("week_str"))
                trend_disp = trend.copy()
                trend_disp["Units"] = trend_disp["Units"].apply(lambda x: f"{float(x):,.0f}")
                trend_disp["Sales"] = trend_disp["Sales"].apply(lambda x: f"${float(x):,.2f}")
                trend_disp = trend_disp.rename(columns={"week_str": "Week"})
                st.dataframe(trend_disp, use_container_width=True, hide_index=True, height=300)

                # Top SKUs
                st.markdown("### Top SKUs (YTD)")
                sku_agg = (wk.groupby("sku", as_index=False)
                             .agg(Units=("Units","sum"), Sales=("Sales","sum"))
                             .sort_values("Units", ascending=False))
                sku_disp = sku_agg.head(25).copy()
                sku_disp["Units"] = sku_disp["Units"].apply(lambda x: f"{float(x):,.0f}")
                sku_disp["Sales"] = sku_disp["Sales"].apply(lambda x: f"${float(x):,.2f}")
                sku_disp = sku_disp.rename(columns={"sku": "SKU"})
                st.dataframe(sku_disp, use_container_width=True, hide_index=True, height=420)

                # Vendor breakdown
                st.markdown("### Vendor breakdown (YTD)")
                v_agg = (wk.groupby("vendor", as_index=False)
                           .agg(Units=("Units","sum"), Sales=("Sales","sum"))
                           .sort_values("Sales", ascending=False))
                v_disp = v_agg.copy()
                v_disp["Units"] = v_disp["Units"].apply(lambda x: f"{float(x):,.0f}")
                v_disp["Sales"] = v_disp["Sales"].apply(lambda x: f"${float(x):,.2f}")
                v_disp = v_disp.rename(columns={"vendor": "Vendor"})
                st.dataframe(v_disp, use_container_width=True, hide_index=True, height=350)

with tab_vendor_scorecard:
    st.subheader("Vendor Scorecard")
    st.caption("Single-vendor view: YTD totals plus WoW and MoM deltas, with top SKUs and retailer breakdown.")

    # Vendors from mapping table (preferred), else from weekly_results->sku join if mapping missing
    try:
        vendors = sorted(pd.read_sql_query("SELECT DISTINCT vendor FROM sku_mapping", conn)["vendor"].dropna().unique().tolist())
    except Exception:
        vendors = []

    if not vendors:
        st.info("No vendors found yet. Upload Vendor-SKU Map and/or data first.")
    else:
        v_sel = st.selectbox("Vendor", vendors, key="vendor_scorecard_sel")

        wk = pd.read_sql_query(
            "SELECT week_start, retailer, sku, units_auto, units_override FROM weekly_results",
            conn,
        )
        wk["week_start"] = pd.to_datetime(wk["week_start"], errors="coerce")
        wk = filter_df_year(wk, year_filter, "week_start")
        wk = wk.dropna(subset=["week_start"])
        wk["retailer"] = wk["retailer"].astype(str).str.strip()
        wk["sku"] = wk["sku"].astype(str).str.strip()

        wk["Units"] = wk["units_override"].where(wk["units_override"].notna(), wk["units_auto"])
        wk["Units"] = pd.to_numeric(wk["Units"], errors="coerce").fillna(0)

        try:
            m = pd.read_sql_query("SELECT retailer, vendor, sku, unit_price FROM sku_mapping", conn)
            m["sku"] = m["sku"].astype(str).str.strip()
            m["retailer"] = m["retailer"].astype(str).str.strip()
            m["vendor"] = m["vendor"].astype(str).str.strip()
        except Exception:
            m = pd.DataFrame(columns=["retailer","vendor","sku","unit_price"])

        if m.empty:
            st.error("sku_mapping table is missing; cannot build vendor scorecard reliably.")
        else:
            wk = wk.merge(m[["retailer","sku","vendor","unit_price"]], on=["retailer","sku"], how="left")
            wk = wk[wk["vendor"] == v_sel].copy()
            wk["unit_price"] = pd.to_numeric(wk["unit_price"], errors="coerce").fillna(0.0)
            wk["Sales"] = wk["Units"] * wk["unit_price"]

            # KPIs
            ytd_units = float(wk["Units"].sum())
            ytd_sales = float(wk["Sales"].sum())

            wk["week_str"] = wk["week_start"].dt.strftime("%Y-%m-%d")
            weeks = sorted(wk["week_str"].unique().tolist())
            if len(weeks) >= 2:
                prev_w, last_w = weeks[-2], weeks[-1]
                prev = wk[wk["week_str"] == prev_w][["Units","Sales"]].sum()
                last = wk[wk["week_str"] == last_w][["Units","Sales"]].sum()
                wow_units = float(last["Units"] - prev["Units"])
                wow_sales = float(last["Sales"] - prev["Sales"])
            else:
                wow_units = 0.0
                wow_sales = 0.0

            if wk["week_start"].notna().any():
                cur_month = wk["week_start"].max().to_period("M")
                prev_month = cur_month - 1
                m_cur = wk[wk["week_start"].dt.to_period("M") == cur_month][["Units","Sales"]].sum()
                m_prev = wk[wk["week_start"].dt.to_period("M") == prev_month][["Units","Sales"]].sum()
                mom_units = float(m_cur["Units"] - m_prev["Units"])
                mom_sales = float(m_cur["Sales"] - m_prev["Sales"])
            else:
                mom_units = 0.0
                mom_sales = 0.0

            k1, k2, k3, k4 = st.columns(4, gap="small")
            k1.metric("Units (YTD)", f"{ytd_units:,.0f}", f"{wow_units:+,.0f} WoW")
            k2.metric("Sales (YTD)", f"${ytd_sales:,.2f}", f"${wow_sales:+,.2f} WoW")
            k3.metric("Units (MoM)", f"{(ytd_units):,.0f}", f"{mom_units:+,.0f} MoM")
            k4.metric("Sales (MoM)", f"${(ytd_sales):,.2f}", f"${mom_sales:+,.2f} MoM")

            st.markdown("### Weekly trend")
            trend = wk.groupby("week_start", as_index=False).agg(Units=("Units","sum"), Sales=("Sales","sum")).sort_values("week_start")
            trend_disp = trend.copy()
            trend_disp['Week'] = trend_disp['week_start'].dt.strftime('%Y-%m-%d')
            trend_disp = trend_disp[['Week','Units','Sales']]
            trend_disp['Units'] = trend_disp['Units'].apply(_fmt_units)
            trend_disp['Sales'] = trend_disp['Sales'].apply(_fmt_money)
            st.dataframe(trend_disp, use_container_width=True, hide_index=True, height=300)
st.markdown("### Top SKUs (YTD)")
            sku_agg = wk.groupby("sku", as_index=False).agg(Units=("Units","sum"), Sales=("Sales","sum")).sort_values("Units", ascending=False)
            left, right = st.columns(2, gap="large")
            with left:
                st.markdown("#### By Units")
                disp = sku_agg[['sku','Units','Sales']].head(25).rename(columns={'sku':'SKU'}).copy()
                disp['Units'] = disp['Units'].apply(_fmt_units)
                disp['Sales'] = disp['Sales'].apply(_fmt_money)
                st.dataframe(disp, use_container_width=True, hide_index=True, height=420)
with right:
                st.markdown("#### By Sales")
                sku_sales = sku_agg.sort_values("Sales", ascending=False)
                disp = sku_sales[['sku','Sales','Units']].head(25).rename(columns={'sku':'SKU'}).copy()
                disp['Units'] = disp['Units'].apply(_fmt_units)
                disp['Sales'] = disp['Sales'].apply(_fmt_money)
                st.dataframe(disp, use_container_width=True, hide_index=True, height=420)
st.markdown("### Retailer breakdown (YTD)")
            r_agg = wk.groupby("retailer", as_index=False).agg(Units=("Units","sum"), Sales=("Sales","sum")).sort_values("Sales", ascending=False)
            r_disp = r_agg.rename(columns={'retailer':'Retailer'}).copy()
            if 'Units' in r_disp.columns: r_disp['Units'] = r_disp['Units'].apply(_fmt_units)
            if 'Sales' in r_disp.columns: r_disp['Sales'] = r_disp['Sales'].apply(_fmt_money)
            st.dataframe(r_disp, use_container_width=True, hide_index=True, height=350)
with tab_vendor_totals:
    st.subheader("Vendor Totals")
    st.caption("Totals across ALL retailers for the selected weeks. Left = Units, Right = Sales ($).")

    label_to_start = {lbl: start.isoformat() for start, _, lbl in week_meta}
    selected_starts = [label_to_start[lbl] for lbl in display_weeks if lbl in label_to_start]

    if not selected_starts:
        st.info("Select at least one week.")
    else:
        placeholders = ",".join(["?"] * len(selected_starts))
        wk = pd.read_sql_query(
            f'''
            SELECT week_start, retailer, sku, units_auto, units_override
            FROM weekly_results
            WHERE week_start IN ({placeholders})
            ''',
            conn,
            params=selected_starts
        )

        wk = filter_df_year(wk, year_filter, "week_start")

        wk = filter_df_year(wk, year_filter, "week_start")

        if wk.empty:
            st.info("No unit data found for the selected weeks yet.")
        else:
            wk["Units"] = wk["units_override"].where(wk["units_override"].notna(), wk["units_auto"])
            wk["Units"] = pd.to_numeric(wk["Units"], errors="coerce").fillna(0)

            mapping_all = pd.read_sql_query(
                '''
                SELECT retailer, vendor, sku, unit_price
                FROM sku_mapping
                WHERE active = 1
                ''',
                conn
            )
            mapping_all["unit_price"] = pd.to_numeric(mapping_all["unit_price"], errors="coerce").fillna(0)

            dfm = wk.merge(mapping_all, on=["retailer", "sku"], how="left")
            dfm["vendor"] = dfm["vendor"].fillna("Unknown")
            dfm["unit_price"] = pd.to_numeric(dfm["unit_price"], errors="coerce").fillna(0)
            dfm["Sales"] = (dfm["Units"] * dfm["unit_price"]).round(2)

            agg = dfm.groupby("vendor", as_index=False).agg(
                Units=("Units", "sum"),
                Sales=("Sales", "sum"),
            )

            # Rankings (descending)
            agg["Units Rank"] = agg["Units"].rank(method="dense", ascending=False).astype(int)
            agg["Sales Rank"] = agg["Sales"].rank(method="dense", ascending=False).astype(int)

            # Sort descending
            agg_units = agg[["vendor","Units Rank","Units"]].sort_values("Units", ascending=False).rename(columns={"vendor":"Vendor"})
            agg_sales = agg[["vendor","Sales Rank","Sales"]].sort_values("Sales", ascending=False).rename(columns={"vendor":"Vendor"}).copy()
            agg_sales["Sales"] = pd.to_numeric(agg_sales["Sales"], errors="coerce").round(2).apply(fmt_currency_str)

            # Total rows
            agg_units = pd.concat([agg_units, pd.DataFrame([{"Vendor":"Total","Units": float(agg_units["Units"].sum())}])], ignore_index=True)
            total_sales = float(pd.to_numeric(agg["Sales"], errors="coerce").sum())
            agg_sales = pd.concat([agg_sales, pd.DataFrame([{"Vendor":"Total","Sales": fmt_currency_str(total_sales)}])], ignore_index=True)

            left_col, right_col = st.columns([1, 1], gap="small")
            with left_col:
                st.markdown("#### Total Units by Vendor")
                st.dataframe(
                    agg_units,
                    use_container_width=True,
                    height=750,
                    hide_index=True,
                    column_config={
                        "Vendor": st.column_config.TextColumn(width="small"),
                        "Units Rank": st.column_config.NumberColumn(format="%.0f", width="small"),
                        "Units": st.column_config.NumberColumn(format="%.0f", width="small"),
                    },
                )

            with right_col:
                st.markdown("#### Total Sales ($) by Vendor")
                st.dataframe(
                    agg_sales,
                    use_container_width=True,
                    height=750,
                    hide_index=True,
                    column_config={
                        "Vendor": st.column_config.TextColumn(width="small"),
                        "Sales Rank": st.column_config.NumberColumn(format="%.0f", width="small"),
                        "Sales": st.column_config.TextColumn(width="small"),
                    },
                )




with tab_health:
    st.subheader("Health Flags")
    st.info("Temporarily disabled to stabilize scorecard summaries. Will be reintroduced cleanly next.")

with tab_backup:

    render_backup_restore_tab(DB_FILE, "Vendor-SKU Map.xlsx")